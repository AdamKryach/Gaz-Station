# -*- coding: utf-8 -*-
"""
Module de Rapports et Statistiques
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, date, timedelta
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.dates as mdates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

class Reports:
    def __init__(self, parent, db_manager):
        self.parent = parent
        self.db_manager = db_manager
        self.setup_interface()
        self.load_dashboard_stats()
    
    def setup_interface(self):
        """Configuration de l'interface des rapports"""
        # Notebook pour les différents rapports
        self.notebook = ttk.Notebook(self.parent)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Onglet Tableau de Bord
        dashboard_frame = ttk.Frame(self.notebook)
        self.notebook.add(dashboard_frame, text="Tableau de Bord")
        self.setup_dashboard(dashboard_frame)
        
        # Onglet Rapports de Ventes
        sales_frame = ttk.Frame(self.notebook)
        self.notebook.add(sales_frame, text="Rapports de Ventes")
        self.setup_sales_reports(sales_frame)
        
        # Onglet Rapports Clients
        clients_frame = ttk.Frame(self.notebook)
        self.notebook.add(clients_frame, text="Rapports Clients")
        self.setup_clients_reports(clients_frame)
        
        # Onglet Rapports Financiers
        financial_frame = ttk.Frame(self.notebook)
        self.notebook.add(financial_frame, text="Rapports Financiers")
        self.setup_financial_reports(financial_frame)
    
    def setup_dashboard(self, parent):
        """Configuration du tableau de bord moderne avec graphiques"""
        # Définir les couleurs modernes pour les graphiques
        self.chart_colors = {
            'primary': '#1976D2',       # Bleu principal
            'secondary': '#FF5722',     # Orange accent
            'success': '#4CAF50',       # Vert succès
            'warning': '#FFC107',       # Jaune avertissement
            'error': '#F44336',         # Rouge erreur
            'background': '#F5F5F5',    # Fond gris clair
            'text': '#212121',          # Texte principal
            'grid': '#E0E0E0'           # Grille de graphique
        }
        
        # Configurer matplotlib pour un style moderne
        plt.style.use('ggplot')
        
        main_frame = ttk.Frame(parent)
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Section statistiques rapides avec cartes modernes
        stats_header = ttk.Frame(main_frame)
        stats_header.pack(fill='x', pady=(0, 15))
        
        # Titre de section avec icône
        ttk.Label(stats_header, 
                 text="📊", 
                 font=('Segoe UI', 18),
                 foreground=self.chart_colors['primary']).pack(side='left', padx=(0, 10))
        
        ttk.Label(stats_header, 
                 text="Statistiques Rapides", 
                 font=('Segoe UI', 16, 'bold'),
                 foreground=self.chart_colors['text']).pack(side='left')
        
        # Bouton actualiser moderne
        refresh_btn = ttk.Button(stats_header, 
                               text="🔄 Actualiser",
                               command=self.load_dashboard_stats,
                               style='Touch.TButton')
        refresh_btn.pack(side='right')
        
        # Conteneur de cartes de statistiques
        stats_container = ttk.Frame(main_frame)
        stats_container.pack(fill='x', pady=(0, 20))
        
        # Variables pour les stats
        self.stats_vars = {}
        
        # Définition des statistiques avec icônes
        stats_info = [
            ('transactions_jour', '🧾 Transactions Aujourd\'hui', '0', 0, 0, '#1976D2'),
            ('ca_jour', '💰 CA Aujourd\'hui (DH)', '0', 0, 1, '#4CAF50'),
            ('litres_jour', '⛽ Litres Vendus Aujourd\'hui', '0', 0, 2, '#FF5722'),
            ('transactions_mois', '📅 Transactions ce Mois', '0', 1, 0, '#1976D2'),
            ('ca_mois', '📈 CA ce Mois (DH)', '0', 1, 1, '#4CAF50'),
            ('litres_mois', '🛢️ Litres Vendus ce Mois', '0', 1, 2, '#FF5722'),
            ('clients_actifs', '👥 Clients Actifs', '0', 2, 0, '#9C27B0'),
            ('factures_impayees', '⚠️ Factures Impayées', '0', 2, 1, '#F44336'),
            ('soldes_positifs', '💲 Soldes Positifs Totaux (DH)', '0', 2, 2, '#009688')
        ]
        
        # Créer une grille pour les cartes de statistiques
        for stat_id, label, default_value, row, col, color in stats_info:
            # Créer une carte moderne pour chaque statistique
            card = ttk.Frame(stats_container, padding=10)
            card.grid(row=row, column=col, padx=10, pady=10, sticky='nsew')
            
            # Titre de la statistique
            ttk.Label(card, 
                     text=label, 
                     font=('Segoe UI', 11),
                     foreground=self.chart_colors['text']).pack(anchor='w')
            
            # Valeur de la statistique
            self.stats_vars[stat_id] = tk.StringVar(value=default_value)
            ttk.Label(card, 
                     textvariable=self.stats_vars[stat_id], 
                     font=('Segoe UI', 22, 'bold'),
                     foreground=color).pack(pady=(5, 0))
        
        # Configuration responsive de la grille
        for i in range(3):
            stats_container.columnconfigure(i, weight=1)
        
        # Section graphiques avec titre moderne
        charts_header = ttk.Frame(main_frame)
        charts_header.pack(fill='x', pady=(20, 15))
        
        # Titre de section avec icône
        ttk.Label(charts_header, 
                 text="📈", 
                 font=('Segoe UI', 18),
                 foreground=self.chart_colors['primary']).pack(side='left', padx=(0, 10))
        
        ttk.Label(charts_header, 
                 text="Graphiques et Analyses", 
                 font=('Segoe UI', 16, 'bold'),
                 foreground=self.chart_colors['text']).pack(side='left')
        
        # Conteneur de graphiques
        charts_frame = ttk.Frame(main_frame)
        charts_frame.pack(fill='both', expand=True)
        
        # Notebook pour les graphiques
        charts_notebook = ttk.Notebook(charts_frame, style='Touch.TNotebook')
        charts_notebook.pack(fill='both', expand=True)
        
        # Graphique Évolution des Ventes
        sales_chart_frame = ttk.Frame(charts_notebook)
        charts_notebook.add(sales_chart_frame, text="Évolution des Ventes")
        self.create_sales_chart(sales_chart_frame)
        
        # Graphique Répartition par Carburant
        fuel_chart_frame = ttk.Frame(charts_notebook)
        charts_notebook.add(fuel_chart_frame, text="Répartition par Carburant")
        self.create_fuel_chart(fuel_chart_frame)
    
    def setup_sales_reports(self, parent):
        """Configuration des rapports de ventes"""
        main_frame = ttk.Frame(parent)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Filtres
        filter_frame = ttk.LabelFrame(main_frame, text="Filtres", padding=15)
        filter_frame.pack(fill='x', pady=(0, 10))
        
        # Ligne 1 - Dates
        date_frame = ttk.Frame(filter_frame)
        date_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(date_frame, text="Du:", style='Touch.TLabel').pack(side='left')
        self.sales_date_from = tk.StringVar(value=datetime.now().strftime("%Y-%m-01"))
        ttk.Entry(date_frame, textvariable=self.sales_date_from, width=15).pack(side='left', padx=(5, 20))
        
        ttk.Label(date_frame, text="Au:", style='Touch.TLabel').pack(side='left')
        self.sales_date_to = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        ttk.Entry(date_frame, textvariable=self.sales_date_to, width=15).pack(side='left', padx=(5, 20))
        
        # Ligne 2 - Station et Carburant
        filter2_frame = ttk.Frame(filter_frame)
        filter2_frame.pack(fill='x')
        
        ttk.Label(filter2_frame, text="Station:", style='Touch.TLabel').pack(side='left')
        self.sales_station_var = tk.StringVar()
        self.sales_station_combo = ttk.Combobox(filter2_frame, textvariable=self.sales_station_var,
                                              state='readonly', width=20)
        self.sales_station_combo.pack(side='left', padx=(5, 20))
        
        ttk.Label(filter2_frame, text="Carburant:", style='Touch.TLabel').pack(side='left')
        self.sales_fuel_var = tk.StringVar()
        self.sales_fuel_combo = ttk.Combobox(filter2_frame, textvariable=self.sales_fuel_var,
                                           state='readonly', width=20)
        self.sales_fuel_combo.pack(side='left', padx=(5, 20))
        
        # Boutons
        btn_frame = ttk.Frame(filter_frame)
        btn_frame.pack(fill='x', pady=(10, 0))
        
        ttk.Button(btn_frame, text="Générer Rapport",
                  command=self.generate_sales_report,
                  style='Touch.TButton').pack(side='left', padx=(0, 10))
        
        ttk.Button(btn_frame, text="Exporter Excel",
                  command=self.export_sales_excel,
                  style='Touch.TButton').pack(side='left')
        
        # Résultats
        results_frame = ttk.LabelFrame(main_frame, text="Résultats", padding=10)
        results_frame.pack(fill='both', expand=True)
        
        # Treeview pour afficher les résultats
        columns = ('Date', 'Station', 'Client', 'Carburant', 'Quantité', 'Prix/L', 'Montant')
        self.sales_tree = ttk.Treeview(results_frame, columns=columns, show='headings',
                                      style='Touch.Treeview')
        
        for col in columns:
            self.sales_tree.heading(col, text=col)
            self.sales_tree.column(col, width=100)
        
        scrollbar_sales = ttk.Scrollbar(results_frame, orient='vertical', command=self.sales_tree.yview)
        self.sales_tree.configure(yscrollcommand=scrollbar_sales.set)
        
        self.sales_tree.pack(side='left', fill='both', expand=True)
        scrollbar_sales.pack(side='right', fill='y')
        
        # Résumé
        summary_sales_frame = ttk.Frame(results_frame)
        summary_sales_frame.pack(fill='x', pady=(10, 0))
        
        self.sales_summary_vars = {
            'total_transactions': tk.StringVar(value="Total Transactions: 0"),
            'total_litres': tk.StringVar(value="Total Litres: 0.0"),
            'total_montant': tk.StringVar(value="Total Montant: 0.00 DH")
        }
        
        for i, (key, var) in enumerate(self.sales_summary_vars.items()):
            ttk.Label(summary_sales_frame, textvariable=var, 
                     font=('Arial', 11, 'bold')).grid(row=0, column=i, padx=20, sticky='w')
        
        # Charger les données initiales
        self.load_sales_filters()
    
    def setup_clients_reports(self, parent):
        """Configuration des rapports clients"""
        main_frame = ttk.Frame(parent)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Types de rapports clients
        report_frame = ttk.LabelFrame(main_frame, text="Types de Rapports", padding=15)
        report_frame.pack(fill='x', pady=(0, 10))
        
        self.client_report_type = tk.StringVar(value="soldes")
        
        reports_types = [
            ("soldes", "Soldes Clients"),
            ("consommation", "Consommation par Client"),
            ("paiements", "Historique Paiements"),
            ("factures", "Factures par Client")
        ]
        
        for value, text in reports_types:
            ttk.Radiobutton(report_frame, text=text, variable=self.client_report_type,
                          value=value).pack(anchor='w', pady=2)
        
        # Boutons
        ttk.Button(report_frame, text="Générer Rapport Client",
                  command=self.generate_client_report,
                  style='Touch.TButton').pack(pady=(10, 0))
        
        # Résultats clients
        client_results_frame = ttk.LabelFrame(main_frame, text="Résultats", padding=10)
        client_results_frame.pack(fill='both', expand=True)
        
        columns_client = ('Client', 'Type', 'Valeur1', 'Valeur2', 'Valeur3')
        self.clients_tree = ttk.Treeview(client_results_frame, columns=columns_client, 
                                        show='headings', style='Touch.Treeview')
        
        for col in columns_client:
            self.clients_tree.heading(col, text=col)
            self.clients_tree.column(col, width=120)
        
        scrollbar_clients = ttk.Scrollbar(client_results_frame, orient='vertical', 
                                         command=self.clients_tree.yview)
        self.clients_tree.configure(yscrollcommand=scrollbar_clients.set)
        
        self.clients_tree.pack(side='left', fill='both', expand=True)
        scrollbar_clients.pack(side='right', fill='y')
    
    def setup_financial_reports(self, parent):
        """Configuration des rapports financiers"""
        main_frame = ttk.Frame(parent)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Types de rapports financiers
        financial_frame = ttk.LabelFrame(main_frame, text="Rapports Financiers", padding=15)
        financial_frame.pack(fill='x', pady=(0, 10))
        
        # Période
        period_frame = ttk.Frame(financial_frame)
        period_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(period_frame, text="Période:", style='Touch.TLabel').pack(side='left')
        self.financial_period = tk.StringVar(value="ce_mois")
        period_combo = ttk.Combobox(period_frame, textvariable=self.financial_period,
                                   values=["cette_semaine", "ce_mois", "trimestre", "annee"],
                                   state='readonly', width=15)
        period_combo.pack(side='left', padx=(5, 20))
        
        # Boutons
        buttons_frame = ttk.Frame(financial_frame)
        buttons_frame.pack(fill='x')
        
        ttk.Button(buttons_frame, text="Chiffre d'Affaires",
                  command=self.generate_ca_report,
                  style='Touch.TButton').pack(side='left', padx=(0, 10))
        
        ttk.Button(buttons_frame, text="Bilan Créances",
                  command=self.generate_credits_report,
                  style='Touch.TButton').pack(side='left', padx=(0, 10))
        
        ttk.Button(buttons_frame, text="État Factures",
                  command=self.generate_invoices_status,
                  style='Touch.TButton').pack(side='left')
        
        # Résultats financiers
        financial_results_frame = ttk.LabelFrame(main_frame, text="Résultats Financiers", padding=10)
        financial_results_frame.pack(fill='both', expand=True)
        
        # Text widget pour afficher les rapports financiers
        self.financial_text = tk.Text(financial_results_frame, wrap='word', font=('Arial', 11))
        financial_scrollbar = ttk.Scrollbar(financial_results_frame, orient='vertical', 
                                           command=self.financial_text.yview)
        self.financial_text.configure(yscrollcommand=financial_scrollbar.set)
        
        self.financial_text.pack(side='left', fill='both', expand=True)
        financial_scrollbar.pack(side='right', fill='y')
    
    def load_dashboard_stats(self):
        """Charger les statistiques du tableau de bord"""
        try:
            today = datetime.now().strftime('%Y-%m-%d')
            first_day_month = datetime.now().strftime('%Y-%m-01')
            
            # Transactions aujourd'hui
            query = "SELECT COUNT(*) FROM transactions WHERE DATE(date_transaction) = DATE(?)"
            result = self.db_manager.execute_query(query, (today,), use_cache=True, cache_timeout=300, table="transactions")
            self.stats_vars['transactions_jour'].set(str(result[0][0] if result else 0))
            
            # CA aujourd'hui
            query = "SELECT COALESCE(SUM(montant_total), 0) FROM transactions WHERE DATE(date_transaction) = DATE(?)"
            result = self.db_manager.execute_query(query, (today,), use_cache=True, cache_timeout=300, table="transactions")
            ca_jour = result[0][0] if result else 0
            self.stats_vars['ca_jour'].set(f"{ca_jour:.2f}")
            
            # Litres vendus aujourd'hui
            query = "SELECT COALESCE(SUM(quantite), 0) FROM transactions WHERE DATE(date_transaction) = DATE(?)"
            result = self.db_manager.execute_query(query, (today,), use_cache=True, cache_timeout=300, table="transactions")
            litres_jour = result[0][0] if result else 0
            self.stats_vars['litres_jour'].set(f"{litres_jour:.1f}")
            
            # Transactions ce mois
            query = "SELECT COUNT(*) FROM transactions WHERE DATE(date_transaction) >= DATE(?)"
            result = self.db_manager.execute_query(query, (first_day_month,), use_cache=True, cache_timeout=300, table="transactions")
            self.stats_vars['transactions_mois'].set(str(result[0][0] if result else 0))
            
            # CA ce mois
            query = "SELECT COALESCE(SUM(montant_total), 0) FROM transactions WHERE DATE(date_transaction) >= DATE(?)"
            result = self.db_manager.execute_query(query, (first_day_month,), use_cache=True, cache_timeout=300, table="transactions")
            ca_mois = result[0][0] if result else 0
            self.stats_vars['ca_mois'].set(f"{ca_mois:.2f}")
            
            # Litres vendus ce mois
            query = "SELECT COALESCE(SUM(quantite), 0) FROM transactions WHERE DATE(date_transaction) >= DATE(?)"
            result = self.db_manager.execute_query(query, (first_day_month,), use_cache=True, cache_timeout=300, table="transactions")
            litres_mois = result[0][0] if result else 0
            self.stats_vars['litres_mois'].set(f"{litres_mois:.1f}")
            
            # Clients actifs
            query = "SELECT COUNT(*) FROM clients WHERE statut = 'actif'"
            result = self.db_manager.execute_query(query, use_cache=True, cache_timeout=600, table="clients")
            self.stats_vars['clients_actifs'].set(str(result[0][0] if result else 0))
            
            # Factures impayées
            query = "SELECT COUNT(*) FROM factures WHERE statut = 'impayee'"
            result = self.db_manager.execute_query(query, use_cache=True, cache_timeout=300, table="factures")
            self.stats_vars['factures_impayees'].set(str(result[0][0] if result else 0))
            
            # Soldes positifs totaux
            query = "SELECT COALESCE(SUM(solde_actuel), 0) FROM clients WHERE solde_actuel > 0"
            result = self.db_manager.execute_query(query, use_cache=True, cache_timeout=600, table="clients")
            soldes_positifs = result[0][0] if result else 0
            self.stats_vars['soldes_positifs'].set(f"{soldes_positifs:.2f}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors du chargement des statistiques: {str(e)}")
    
    def create_sales_chart(self, parent):
        """Créer le graphique d'évolution des ventes"""
        try:
            # Récupérer les données des 7 derniers jours
            dates = []
            amounts = []
            
            for i in range(7):
                date_obj = datetime.now() - timedelta(days=6-i)
                date_str = date_obj.strftime('%Y-%m-%d')
                dates.append(date_obj)
                
                query = "SELECT COALESCE(SUM(montant_total), 0) FROM transactions WHERE DATE(date_transaction) = DATE(?)"
                result = self.db_manager.execute_query(query, (date_str,), use_cache=True, cache_timeout=300, table="transactions")
                amounts.append(result[0][0] if result else 0)
            
            # Créer le graphique (très agrandi pour écran tactile)
            fig, ax = plt.subplots(figsize=(16, 10))
            ax.plot(dates, amounts, marker='o', linewidth=2, markersize=6)
            ax.set_title('Évolution des Ventes (7 derniers jours)', fontsize=14, fontweight='bold')
            ax.set_xlabel('Date')
            ax.set_ylabel('Montant (DH)')
            ax.grid(True, alpha=0.3)
            
            # Format des dates
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
            ax.xaxis.set_major_locator(mdates.DayLocator())
            plt.xticks(rotation=45)
            
            plt.tight_layout()
            
            # Intégrer dans tkinter
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)
            
        except Exception as e:
            ttk.Label(parent, text=f"Erreur graphique: {str(e)}", 
                     foreground='red').pack(expand=True)
    
    def create_fuel_chart(self, parent):
        """Créer le graphique de répartition par carburant"""
        try:
            # Récupérer les données
            query = """
                SELECT c.nom, COALESCE(SUM(t.montant_total), 0) as total
                FROM carburants c
                LEFT JOIN transactions t ON c.id = t.carburant_id 
                    AND DATE(t.date_transaction) >= DATE('now', '-30 days')
                GROUP BY c.id, c.nom
                HAVING total > 0
                ORDER BY total DESC
            """
            
            results = self.db_manager.execute_query(query, use_cache=True, cache_timeout=300, table=["carburants", "transactions"])
            
            if results:
                labels = [row[0] for row in results]
                sizes = [row[1] for row in results]
                
                # Créer une figure moderne avec deux graphiques côte à côte
                fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 12), facecolor=self.chart_colors['background'])
                
                # Utiliser les couleurs du thème
                colors = [self.chart_colors['primary'], self.chart_colors['secondary'], 
                          self.chart_colors['success'], self.chart_colors['warning'], 
                          self.chart_colors['error']]
                
                # Graphique en donut moderne
                wedges, texts, autotexts = ax1.pie(sizes, 
                                                 labels=None,  # Pas de labels sur le graphique
                                                 autopct='%1.1f%%',
                                                 pctdistance=0.85,
                                                 colors=colors[:len(labels)], 
                                                 startangle=90,
                                                 wedgeprops=dict(width=0.5, edgecolor='white'),
                                                 textprops=dict(color='white', fontweight='bold'))
                
                # Ajouter un cercle au milieu pour créer un effet donut
                centre_circle = plt.Circle((0, 0), 0.35, fc='white')
                ax1.add_patch(centre_circle)
                
                # Légende personnalisée
                ax1.legend(wedges, labels, 
                         title="Types de Carburant",
                         loc="center left",
                         bbox_to_anchor=(0.9, 0, 0.5, 1),
                         frameon=False)
                
                ax1.set_title('Répartition des Ventes par Carburant\n(30 derniers jours)', 
                           fontsize=14, fontweight='bold', color=self.chart_colors['text'])
                ax1.axis('equal')
                
                # Graphique à barres pour les montants
                bars = ax2.bar(labels, sizes, color=colors[:len(labels)],
                             width=0.6, edgecolor='white', linewidth=1)
                
                # Ajouter les valeurs au-dessus des barres
                for bar in bars:
                    height = bar.get_height()
                    ax2.annotate(f'{height:,.2f} DH',
                               xy=(bar.get_x() + bar.get_width() / 2, height),
                               xytext=(0, 3),
                               textcoords="offset points",
                               ha='center', va='bottom',
                               fontsize=10, fontweight='bold',
                               color=self.chart_colors['text'])
                
                # Configurer les axes du graphique à barres
                ax2.set_title('Montants par Type de Carburant', 
                             fontsize=14, fontweight='bold', color=self.chart_colors['text'])
                ax2.spines['top'].set_visible(False)
                ax2.spines['right'].set_visible(False)
                ax2.spines['left'].set_color(self.chart_colors['grid'])
                ax2.spines['bottom'].set_color(self.chart_colors['grid'])
                ax2.set_ylabel('Montant (DH)', fontsize=12, color=self.chart_colors['text'])
                ax2.grid(axis='y', linestyle='--', alpha=0.7, color=self.chart_colors['grid'])
                plt.xticks(color=self.chart_colors['text'])
                plt.yticks(color=self.chart_colors['text'])
                
                plt.tight_layout()
                
                # Intégrer dans tkinter avec un cadre moderne
                chart_frame = ttk.Frame(parent, padding=10)
                chart_frame.pack(fill='both', expand=True)
                
                canvas = FigureCanvasTkAgg(fig, chart_frame)
                canvas.draw()
                canvas.get_tk_widget().pack(fill='both', expand=True, padx=5, pady=5)
            else:
                ttk.Label(parent, text="Aucune donnée à afficher", 
                         font=('Arial', 12), foreground=self.chart_colors['text']).pack(expand=True)
                
        except Exception as e:
            ttk.Label(parent, text=f"Erreur graphique: {str(e)}", 
                     foreground=self.chart_colors['error']).pack(expand=True)
    
    def load_sales_filters(self):
        """Charger les filtres pour les rapports de ventes"""
        try:
            # Charger les stations
            stations = self.db_manager.execute_query("SELECT id, nom FROM stations ORDER BY nom", use_cache=True, cache_timeout=600, table="stations")
            station_list = ["Toutes"] + [f"{station[0]} - {station[1]}" for station in stations]
            self.sales_station_combo['values'] = station_list
            self.sales_station_combo.current(0)
            
            # Charger les carburants
            fuels = self.db_manager.execute_query("SELECT id, nom FROM carburants ORDER BY nom", use_cache=True, cache_timeout=600, table="carburants")
            fuel_list = ["Tous"] + [f"{fuel[0]} - {fuel[1]}" for fuel in fuels]
            self.sales_fuel_combo['values'] = fuel_list
            self.sales_fuel_combo.current(0)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors du chargement des filtres: {str(e)}")
    
    def generate_sales_report(self):
        """Générer le rapport de ventes"""
        try:
            # Vider les résultats précédents
            for item in self.sales_tree.get_children():
                self.sales_tree.delete(item)
            
            # Construire la requête
            where_conditions = []
            params = []
            
            # Filtres de date
            date_from = self.sales_date_from.get()
            date_to = self.sales_date_to.get()
            
            if date_from:
                where_conditions.append("DATE(t.date_transaction) >= DATE(?)")
                params.append(date_from)
            
            if date_to:
                where_conditions.append("DATE(t.date_transaction) <= DATE(?)")
                params.append(date_to)
            
            # Filtre station
            station_filter = self.sales_station_var.get()
            if station_filter and station_filter != "Toutes" and " - " in station_filter:
                station_id = int(station_filter.split(" - ")[0])
                where_conditions.append("t.station_id = ?")
                params.append(station_id)
            
            # Filtre carburant
            fuel_filter = self.sales_fuel_var.get()
            if fuel_filter and fuel_filter != "Tous" and " - " in fuel_filter:
                fuel_id = int(fuel_filter.split(" - ")[0])
                where_conditions.append("t.carburant_id = ?")
                params.append(fuel_id)
            
            where_clause = " AND ".join(where_conditions)
            if where_clause:
                where_clause = "WHERE " + where_clause
            
            query = f"""
                SELECT 
                    DATE(t.date_transaction) as date,
                    s.nom as station,
                    CASE 
                        WHEN c.type_client = 'entreprise' AND c.entreprise IS NOT NULL 
                        THEN c.entreprise 
                        ELSE c.nom || ' ' || COALESCE(c.prenom, '')
                    END as client,
                    car.nom as carburant,
                    t.quantite,
                    t.prix_unitaire,
                    t.montant_total
                FROM transactions t
                JOIN stations s ON t.station_id = s.id
                JOIN clients c ON t.client_id = c.id
                JOIN carburants car ON t.carburant_id = car.id
                {where_clause}
                ORDER BY t.date_transaction DESC, t.id DESC
            """
            
            results = self.db_manager.execute_query(query, params, use_cache=True, cache_timeout=300, table=["transactions", "stations", "clients", "carburants"])
            
            total_transactions = len(results)
            total_litres = 0
            total_montant = 0
            
            for row in results:
                # Afficher dans le treeview
                values = (
                    row[0],  # Date
                    row[1],  # Station
                    row[2],  # Client
                    row[3],  # Carburant
                    f"{row[4]:.1f}L",  # Quantité
                    f"{row[5]:.2f}",   # Prix/L
                    f"{row[6]:.2f} DH" # Montant
                )
                self.sales_tree.insert('', 'end', values=values)
                
                # Calculs pour le résumé
                total_litres += row[4]
                total_montant += row[6]
            
            # Mettre à jour le résumé
            self.sales_summary_vars['total_transactions'].set(f"Total Transactions: {total_transactions}")
            self.sales_summary_vars['total_litres'].set(f"Total Litres: {total_litres:.1f}")
            self.sales_summary_vars['total_montant'].set(f"Total Montant: {total_montant:.2f} DH")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération du rapport: {str(e)}")
    
    def export_sales_excel(self):
        """Exporter le rapport de ventes vers Excel"""
        try:
            if not self.sales_tree.get_children():
                messagebox.showwarning("Attention", "Aucune donnée à exporter")
                return
            
            # Demander où sauvegarder
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Exporter le rapport de ventes"
            )
            
            if not filename:
                return
            
            # Créer le workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Rapport de Ventes"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
            
            # En-têtes
            headers = ['Date', 'Station', 'Client', 'Carburant', 'Quantité', 'Prix/L', 'Montant']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
            
            # Données
            for row_idx, item in enumerate(self.sales_tree.get_children(), 2):
                values = self.sales_tree.item(item)['values']
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                    cell.border = border
            
            # Ajuster la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Ajouter le résumé
            summary_row = len(self.sales_tree.get_children()) + 3
            ws.cell(row=summary_row, column=1, value="RÉSUMÉ:").font = Font(bold=True)
            
            for i, (key, var) in enumerate(self.sales_summary_vars.items()):
                ws.cell(row=summary_row + 1 + i, column=1, value=var.get()).font = Font(bold=True)
            
            wb.save(filename)
            messagebox.showinfo("Succès", f"Rapport exporté vers: {filename}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'export: {str(e)}")
    
    def generate_client_report(self):
        """Générer le rapport client sélectionné"""
        try:
            # Vider les résultats précédents
            for item in self.clients_tree.get_children():
                self.clients_tree.delete(item)
            
            report_type = self.client_report_type.get()
            
            if report_type == "soldes":
                self.generate_client_balances_report()
            elif report_type == "consommation":
                self.generate_client_consumption_report()
            elif report_type == "paiements":
                self.generate_client_payments_report()
            elif report_type == "factures":
                self.generate_client_invoices_report()
                
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération du rapport client: {str(e)}")
    
    def generate_client_balances_report(self):
        """Générer le rapport des soldes clients"""
        # Configurer les en-têtes
        self.clients_tree.heading('Client', text='Client')
        self.clients_tree.heading('Type', text='Type')
        self.clients_tree.heading('Valeur1', text='Solde Actuel (DH)')
        self.clients_tree.heading('Valeur2', text='Limite Crédit (DH)')
        self.clients_tree.heading('Valeur3', text='Statut')
        
        query = """
            SELECT 
                CASE 
                    WHEN c.type_client = 'entreprise' AND c.entreprise IS NOT NULL 
                    THEN c.entreprise 
                    ELSE c.nom || ' ' || COALESCE(c.prenom, '')
                END as client,
                c.type_client,
                c.solde_actuel,
                c.credit_limite,
                c.statut
            FROM clients c
            WHERE c.statut = 'actif'
            ORDER BY c.solde_actuel DESC
        """
        
        results = self.db_manager.execute_query(query, use_cache=True, cache_timeout=300, table="clients")
        
        for row in results:
            values = (
                row[0],  # Client
                row[1],  # Type
                f"{row[2]:.2f}",  # Solde
                f"{row[3]:.2f}",  # Limite
                row[4]   # Statut
            )
            self.clients_tree.insert('', 'end', values=values)
    
    def generate_client_consumption_report(self):
        """Générer le rapport de consommation par client"""
        # Configurer les en-têtes
        self.clients_tree.heading('Client', text='Client')
        self.clients_tree.heading('Type', text='Nb Transactions')
        self.clients_tree.heading('Valeur1', text='Total Litres')
        self.clients_tree.heading('Valeur2', text='Montant Total (DH)')
        self.clients_tree.heading('Valeur3', text='Dernière Transaction')
        
        query = """
            SELECT 
                CASE 
                    WHEN c.type_client = 'entreprise' AND c.entreprise IS NOT NULL 
                    THEN c.entreprise 
                    ELSE c.nom || ' ' || COALESCE(c.prenom, '')
                END as client,
                COUNT(t.id) as nb_transactions,
                COALESCE(SUM(t.quantite), 0) as total_litres,
                COALESCE(SUM(t.montant_total), 0) as total_montant,
                MAX(DATE(t.date_transaction)) as derniere_transaction
            FROM clients c
            LEFT JOIN transactions t ON c.id = t.client_id
            WHERE c.statut = 'actif'
            GROUP BY c.id, client
            ORDER BY total_montant DESC
        """
        
        results = self.db_manager.execute_query(query, use_cache=True, cache_timeout=300, table=["clients", "transactions"])
        
        for row in results:
            values = (
                row[0],  # Client
                str(row[1]),  # Nb transactions
                f"{row[2]:.1f}L",  # Total litres
                f"{row[3]:.2f}",   # Montant total
                row[4] or '-'  # Dernière transaction
            )
            self.clients_tree.insert('', 'end', values=values)
    
    def generate_client_payments_report(self):
        """Générer le rapport des paiements clients"""
        # Configurer les en-têtes
        self.clients_tree.heading('Client', text='Client')
        self.clients_tree.heading('Type', text='Nb Paiements')
        self.clients_tree.heading('Valeur1', text='Total Paiements (DH)')
        self.clients_tree.heading('Valeur2', text='Paiements Actifs (DH)')
        self.clients_tree.heading('Valeur3', text='Dernier Paiement')
        
        query = """
            SELECT 
                CASE 
                    WHEN c.type_client = 'entreprise' AND c.entreprise IS NOT NULL 
                    THEN c.entreprise 
                    ELSE c.nom || ' ' || COALESCE(c.prenom, '')
                END as client,
                COUNT(p.id) as nb_paiements,
                COALESCE(SUM(p.montant), 0) as total_paiements,
                COALESCE(SUM(CASE WHEN p.statut = 'actif' THEN p.montant ELSE 0 END), 0) as paiements_actifs,
                MAX(DATE(p.date_paiement)) as dernier_paiement
            FROM clients c
            LEFT JOIN paiements_avance p ON c.id = p.client_id
            WHERE c.statut = 'actif'
            GROUP BY c.id, client
            HAVING nb_paiements > 0
            ORDER BY total_paiements DESC
        """
        
        results = self.db_manager.execute_query(query, use_cache=True, cache_timeout=300, table=["clients", "paiements_avance"])
        
        for row in results:
            values = (
                row[0],  # Client
                str(row[1]),  # Nb paiements
                f"{row[2]:.2f}",  # Total paiements
                f"{row[3]:.2f}",  # Paiements actifs
                row[4] or '-'  # Dernier paiement
            )
            self.clients_tree.insert('', 'end', values=values)
    
    def generate_client_invoices_report(self):
        """Générer le rapport des factures par client"""
        # Configurer les en-têtes
        self.clients_tree.heading('Client', text='Client')
        self.clients_tree.heading('Type', text='Nb Factures')
        self.clients_tree.heading('Valeur1', text='Total HT (DH)')
        self.clients_tree.heading('Valeur2', text='Total TTC (DH)')
        self.clients_tree.heading('Valeur3', text='Factures Impayées')
        
        query = """
            SELECT 
                CASE 
                    WHEN c.type_client = 'entreprise' AND c.entreprise IS NOT NULL 
                    THEN c.entreprise 
                    ELSE c.nom || ' ' || COALESCE(c.prenom, '')
                END as client,
                COUNT(f.id) as nb_factures,
                COALESCE(SUM(f.montant_ht), 0) as total_ht,
                COALESCE(SUM(f.montant_ttc), 0) as total_ttc,
                COUNT(CASE WHEN f.statut = 'impayee' THEN 1 END) as factures_impayees
            FROM clients c
            LEFT JOIN factures f ON c.id = f.client_id
            WHERE c.statut = 'actif'
            GROUP BY c.id, client
            HAVING nb_factures > 0
            ORDER BY total_ttc DESC
        """
        
        results = self.db_manager.execute_query(query, use_cache=True, cache_timeout=300, table=["clients", "factures"])
        
        for row in results:
            values = (
                row[0],  # Client
                str(row[1]),  # Nb factures
                f"{row[2]:.2f}",  # Total HT
                f"{row[3]:.2f}",  # Total TTC
                str(row[4])  # Factures impayées
            )
            self.clients_tree.insert('', 'end', values=values)
    
    def generate_ca_report(self):
        """Générer le rapport de chiffre d'affaires"""
        try:
            period = self.financial_period.get()
            
            # Déterminer la période
            if period == "cette_semaine":
                date_condition = "DATE(t.date_transaction) >= DATE('now', '-7 days')"
                period_label = "cette semaine"
            elif period == "ce_mois":
                date_condition = "DATE(t.date_transaction) >= DATE('now', 'start of month')"
                period_label = "ce mois"
            elif period == "trimestre":
                date_condition = "DATE(t.date_transaction) >= DATE('now', '-3 months')"
                period_label = "ce trimestre"
            else:  # année
                date_condition = "DATE(t.date_transaction) >= DATE('now', 'start of year')"
                period_label = "cette année"
            
            # Requête principale
            query = f"""
                SELECT 
                    s.nom as station,
                    COUNT(t.id) as nb_transactions,
                    COALESCE(SUM(t.quantite), 0) as total_litres,
                    COALESCE(SUM(t.montant_total), 0) as total_ca,
                    COALESCE(AVG(t.montant_total), 0) as ca_moyen
                FROM stations s
                LEFT JOIN transactions t ON s.id = t.station_id AND {date_condition}
                GROUP BY s.id, s.nom
                ORDER BY total_ca DESC
            """
            
            results = self.db_manager.execute_query(query, use_cache=True, cache_timeout=300, table=["stations", "transactions"])
            
            # Construire le rapport
            report = f"RAPPORT DE CHIFFRE D'AFFAIRES - {period_label.upper()}\n"
            report += "=" * 60 + "\n\n"
            
            total_global_ca = 0
            total_global_transactions = 0
            total_global_litres = 0
            
            for row in results:
                station, nb_trans, litres, ca, ca_moyen = row
                total_global_ca += ca
                total_global_transactions += nb_trans
                total_global_litres += litres
                
                report += f"Station: {station}\n"
                report += f"  Transactions: {nb_trans}\n"
                report += f"  Litres vendus: {litres:.1f}L\n"
                report += f"  Chiffre d'affaires: {ca:.2f} DH\n"
                report += f"  CA moyen par transaction: {ca_moyen:.2f} DH\n"
                report += "-" * 40 + "\n\n"
            
            # Total global
            report += "TOTAL GLOBAL:\n"
            report += f"  Total Transactions: {total_global_transactions}\n"
            report += f"  Total Litres: {total_global_litres:.1f}L\n"
            report += f"  Total CA: {total_global_ca:.2f} DH\n"
            
            if total_global_transactions > 0:
                ca_moyen_global = total_global_ca / total_global_transactions
                report += f"  CA moyen par transaction: {ca_moyen_global:.2f} DH\n"
            
            # Afficher le rapport
            self.financial_text.delete('1.0', tk.END)
            self.financial_text.insert('1.0', report)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération du rapport CA: {str(e)}")
    
    def generate_credits_report(self):
        """Générer le bilan des créances"""
        try:
            query = """
                SELECT 
                    CASE 
                        WHEN c.type_client = 'entreprise' AND c.entreprise IS NOT NULL 
                        THEN c.entreprise 
                        ELSE c.nom || ' ' || COALESCE(c.prenom, '')
                    END as client,
                    c.solde_actuel,
                    c.credit_limite,
                    COUNT(t.id) as nb_transactions_credit,
                    COALESCE(SUM(CASE WHEN t.type_paiement = 'credit' THEN t.montant_total ELSE 0 END), 0) as total_credit,
                    MAX(DATE(t.date_transaction)) as derniere_transaction
                FROM clients c
                LEFT JOIN transactions t ON c.id = t.client_id AND t.type_paiement = 'credit'
                WHERE c.statut = 'actif' AND (c.solde_actuel != 0 OR COUNT(t.id) > 0)
                GROUP BY c.id, client, c.solde_actuel, c.credit_limite
                ORDER BY c.solde_actuel DESC
            """
            
            results = self.db_manager.execute_query(query, use_cache=True, cache_timeout=300, table=["clients", "transactions"])
            
            report = "BILAN DES CRÉANCES\n"
            report += "=" * 50 + "\n\n"
            
            total_creances_positives = 0
            total_creances_negatives = 0
            nb_clients_debiteurs = 0
            nb_clients_crediteurs = 0
            
            report += "DÉTAIL PAR CLIENT:\n"
            report += "-" * 30 + "\n\n"
            
            for row in results:
                client, solde, limite, nb_trans, total_credit, derniere = row
                
                if solde > 0:
                    total_creances_positives += solde
                    nb_clients_crediteurs += 1
                elif solde < 0:
                    total_creances_negatives += abs(solde)
                    nb_clients_debiteurs += 1
                
                report += f"Client: {client}\n"
                report += f"  Solde actuel: {solde:.2f} DH\n"
                report += f"  Limite de crédit: {limite:.2f} DH\n"
                report += f"  Transactions à crédit: {nb_trans}\n"
                report += f"  Total vendu à crédit: {total_credit:.2f} DH\n"
                report += f"  Dernière transaction: {derniere or 'Aucune'}\n"
                report += "-" * 30 + "\n\n"
            
            # Résumé
            report += "RÉSUMÉ:\n"
            report += f"  Clients créditeurs (solde positif): {nb_clients_crediteurs}\n"
            report += f"  Total créances positives: {total_creances_positives:.2f} DH\n"
            report += f"  Clients débiteurs (solde négatif): {nb_clients_debiteurs}\n"
            report += f"  Total créances négatives: {total_creances_negatives:.2f} DH\n"
            report += f"  Solde net: {(total_creances_positives - total_creances_negatives):.2f} DH\n"
            
            # Afficher le rapport
            self.financial_text.delete('1.0', tk.END)
            self.financial_text.insert('1.0', report)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération du bilan: {str(e)}")
    
    def generate_invoices_status(self):
        """Générer l'état des factures"""
        try:
            query = """
                SELECT 
                    f.statut,
                    COUNT(f.id) as nb_factures,
                    COALESCE(SUM(f.montant_ht), 0) as total_ht,
                    COALESCE(SUM(f.montant_ttc), 0) as total_ttc
                FROM factures f
                GROUP BY f.statut
                ORDER BY 
                    CASE f.statut 
                        WHEN 'impayee' THEN 1 
                        WHEN 'payee' THEN 2 
                        WHEN 'annulee' THEN 3 
                        ELSE 4 
                    END
            """
            
            results = self.db_manager.execute_query(query, use_cache=True, cache_timeout=300, table="factures")
            
            report = "ÉTAT DES FACTURES\n"
            report += "=" * 40 + "\n\n"
            
            total_factures = 0
            total_ht_global = 0
            total_ttc_global = 0
            
            for row in results:
                statut, nb, ht, ttc = row
                total_factures += nb
                total_ht_global += ht
                total_ttc_global += ttc
                
                report += f"Statut: {statut.upper()}\n"
                report += f"  Nombre de factures: {nb}\n"
                report += f"  Total HT: {ht:.2f} DH\n"
                report += f"  Total TTC: {ttc:.2f} DH\n"
                report += "-" * 30 + "\n\n"
            
            # Factures récentes
            recent_query = """
                SELECT 
                    f.numero_facture,
                    f.date_facture,
                    CASE 
                        WHEN c.type_client = 'entreprise' AND c.entreprise IS NOT NULL 
                        THEN c.entreprise 
                        ELSE c.nom || ' ' || COALESCE(c.prenom, '')
                    END as client,
                    f.montant_ttc,
                    f.statut
                FROM factures f
                JOIN clients c ON f.client_id = c.id
                WHERE f.statut = 'impayee'
                ORDER BY f.date_facture DESC
                LIMIT 10
            """
            
            recent_results = self.db_manager.execute_query(recent_query, use_cache=True, cache_timeout=300, table=["factures", "clients"])
            
            report += "FACTURES IMPAYÉES RÉCENTES (10 dernières):\n"
            report += "-" * 40 + "\n\n"
            
            for row in recent_results:
                numero, date_fact, client, montant, statut = row
                report += f"N° {numero} - {date_fact}\n"
                report += f"  Client: {client}\n"
                report += f"  Montant: {montant:.2f} DH\n"
                report += f"  Statut: {statut}\n\n"
            
            # Total global
            report += f"TOTAL GÉNÉRAL:\n"
            report += f"  Total factures: {total_factures}\n"
            report += f"  Total HT: {total_ht_global:.2f} DH\n"
            report += f"  Total TTC: {total_ttc_global:.2f} DH\n"
            
            # Afficher le rapport
            self.financial_text.delete('1.0', tk.END)
            self.financial_text.insert('1.0', report)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération de l'état des factures: {str(e)}")
